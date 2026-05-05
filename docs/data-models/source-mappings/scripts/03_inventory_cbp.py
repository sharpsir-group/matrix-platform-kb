#!/usr/bin/env python3
"""
Inventory the SIR business-practice listing checklists (XLSX).

Reads:
  ../../../../raw/current-business-practice/Listing procedure check list January 2024.xlsx
  ../../../../raw/current-business-practice/Check-list Listing Residential Properties and Land ENG 2025-2026.xlsx

Writes:
  raw/cbp_inventory.csv

Phase: Inventory (mechanical, deterministic).
Owner: scripts/03_inventory_cbp.py - do NOT hand-edit raw/cbp_inventory.csv.

Determinism:
  - Sorted by (checklist, sheet, row_no, col_no).
  - csv.QUOTE_MINIMAL.

Heuristic:
  - Each non-empty cell in column A is a step/field/section.
  - kind=field if the cell ends with ':' (or ': *', ': √') - i.e. it
    is a fillable label.
  - kind=section if it is a short title-cased phrase that does not end
    with ':' and is followed by indented field rows (heuristic: text
    such as "Seller Details", "Selling Details", "Property Location").
  - kind=step for procedural rows ("Get title deeds & permits from
    the seller", arrows, etc.).
  - kind=enum_value for cells in the "filters" sheet (lookup options).
"""

from __future__ import annotations

import csv
import re
import sys
import warnings
from pathlib import Path

import openpyxl
from openpyxl.utils.exceptions import InvalidFileException

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

KB_ROOT = Path(__file__).resolve().parent.parent
REPO_ROOT = KB_ROOT.parent.parent.parent
CBP_DIR = REPO_ROOT / "raw" / "current-business-practice"
OUT_CSV = KB_ROOT / "raw" / "cbp_inventory.csv"

CHECKLISTS = {
    "Listing procedure check list January 2024.xlsx": "ListingProcedure_2024",
    "Check-list Listing Residential Properties and Land ENG 2025-2026.xlsx": "ListingChecklist_2025_2026",
}

LABEL_RE = re.compile(r":\s*[√v*✓✔]?\s*$")
ARROW_RE = re.compile(r"^[\u2193\u21d3\u2b07]+$")  # ↓ ⇓ ⬇
SECTION_HEADERS = {
    "Seller Details",
    "Selling Details",
    "Property Location",
    "Apartment Fields",
    "House Fields",
    "Land Fields",
    "Construction & Development",
    "Listing Process Checklist Brokers",
    "Listing Process Checklist Marketing (only after listing agreement is signed)",
    "Property Subtype",
    "Filters",
}


def normalize(text) -> str:
    if text is None:
        return ""
    return " ".join(str(text).split())


def has_required_marker(text: str) -> bool:
    stripped = text.rstrip()
    return any(stripped.endswith(m) for m in ("*", "√", "✓", "✔"))


def field_type_hint(label: str) -> str:
    lower = label.lower()
    if "(yes/no)" in lower or "y/n" in lower:
        return "boolean"
    if any(k in lower for k in ("price", "amount", "fee", "rate", "%", "(sq.m", "size", "area")):
        return "number"
    if "date" in lower or "year" in lower:
        return "date"
    if "(" in label and ")" in label and (",") in label[label.find("(") :]:
        return "enum"
    return "text"


def main() -> int:
    if not CBP_DIR.is_dir():
        print(f"ERROR: CBP source dir not found: {CBP_DIR}", file=sys.stderr)
        return 2

    rows: list[dict] = []
    seen: set[tuple] = set()

    for filename, checklist_name in sorted(CHECKLISTS.items()):
        path = CBP_DIR / filename
        if not path.is_file():
            print(f"ERROR: missing checklist: {path}", file=sys.stderr)
            return 2
        try:
            wb = openpyxl.load_workbook(str(path), data_only=True)
        except InvalidFileException as exc:
            print(f"ERROR: could not open {path}: {exc}", file=sys.stderr)
            return 2

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            current_section = sheet_name  # default section is the sheet name

            for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if not row:
                    continue

                col_a = normalize(row[0]) if len(row) > 0 else ""
                col_b = normalize(row[1]) if len(row) > 1 else ""
                col_c = normalize(row[2]) if len(row) > 2 else ""

                # Skip pure-arrow rows.
                if col_a and ARROW_RE.match(col_a):
                    continue

                if not col_a:
                    continue

                # Update section when we see a known header.
                if col_a in SECTION_HEADERS:
                    current_section = col_a
                    kind = "section"
                elif sheet_name.lower() == "filters":
                    # Lookup-value table; treat header columns as sections,
                    # other cells as enum values.
                    if r_idx == 1:
                        current_section = col_a
                        kind = "section"
                    else:
                        kind = "enum_value"
                elif LABEL_RE.search(col_a):
                    kind = "field"
                elif col_a.endswith(")") and "(" in col_a and len(col_a) > 30:
                    # Long parenthetical = likely a procedural step or field with hint.
                    kind = "field" if ":" in col_a else "step"
                else:
                    kind = "step"

                # Build the canonical label (strip the trailing ":" / marker for fields).
                if kind == "field":
                    field_label = re.sub(r":\s*[√v*✓✔]?\s*$", "", col_a).strip()
                else:
                    field_label = col_a

                # Walk every cell that's also non-empty in this row to capture
                # sub-fields (e.g. enum value rows in 'filters' sheet have
                # values across multiple columns).
                cells_to_record = [(0, col_a)]
                if sheet_name.lower() == "filters" and r_idx >= 2:
                    cells_to_record = [
                        (i, normalize(v))
                        for i, v in enumerate(row)
                        if v not in (None, "")
                    ]

                for col_idx, cell_text in cells_to_record:
                    if not cell_text or ARROW_RE.match(cell_text):
                        continue
                    if sheet_name.lower() == "filters" and r_idx >= 2:
                        # In filters sheet: r_idx >=2 cells are enum values
                        # except for the values-cell that equals col_a header.
                        rec_kind = "enum_value"
                        rec_label = cell_text
                    elif col_idx == 0:
                        rec_kind = kind
                        rec_label = field_label
                    else:
                        # Subsequent columns on field/step rows are example
                        # values or hints; capture as enum_value if short and
                        # alphanumeric.
                        if len(cell_text) > 60 or any(
                            c in cell_text for c in (":", "↓", "@")
                        ):
                            continue
                        rec_kind = "example_value"
                        rec_label = cell_text

                    key = (
                        checklist_name,
                        sheet_name,
                        current_section,
                        rec_kind,
                        rec_label,
                    )
                    if key in seen:
                        continue
                    seen.add(key)

                    rows.append(
                        {
                            "checklist": checklist_name,
                            "sheet": sheet_name,
                            "section": current_section,
                            "kind": rec_kind,
                            "step_label": rec_label,
                            "field_type_hint": field_type_hint(rec_label) if rec_kind == "field" else "",
                            "required_yn": "yes" if rec_kind == "field" and has_required_marker(col_a) else "",
                            "example_value": col_b if rec_kind == "field" and col_idx == 0 else "",
                            "timing_or_note": col_c if rec_kind in ("field", "step") and col_idx == 0 else "",
                            "row_no": r_idx,
                            "col_no": col_idx,
                        }
                    )

    rows.sort(
        key=lambda r: (
            r["checklist"],
            r["sheet"],
            r["section"],
            r["kind"],
            r["row_no"],
            r["col_no"],
            r["step_label"],
        )
    )

    OUT_CSV.parent.mkdir(parents=True, exist_ok=True)
    with OUT_CSV.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(
            fh,
            fieldnames=[
                "checklist",
                "sheet",
                "section",
                "kind",
                "step_label",
                "field_type_hint",
                "required_yn",
                "example_value",
                "timing_or_note",
                "row_no",
                "col_no",
            ],
            quoting=csv.QUOTE_MINIMAL,
        )
        writer.writeheader()
        writer.writerows(rows)

    field_count = sum(1 for r in rows if r["kind"] == "field")
    step_count = sum(1 for r in rows if r["kind"] == "step")
    enum_count = sum(1 for r in rows if r["kind"] == "enum_value")
    print(
        f"[03] wrote {OUT_CSV.relative_to(KB_ROOT)}: {len(rows)} rows "
        f"({field_count} fields, {step_count} steps, {enum_count} enum values)"
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
